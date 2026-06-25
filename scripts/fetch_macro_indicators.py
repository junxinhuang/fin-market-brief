#!/usr/bin/env python3
"""
Fetch first-pass historical data for data/macro-indicators.yaml.

The output is intentionally evidence-oriented: each indicator gets either a
history series or an explicit gap/error. This keeps the dashboard from silently
substituting weak proxies for missing macro data.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import math
import os
import re
import subprocess
import sys
import time
import warnings
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlencode

try:
    import yaml
except ImportError:
    for minor in range(9, 13):
        user_site = Path.home() / "Library" / "Python" / f"3.{minor}" / "lib" / "python" / "site-packages"
        if user_site.exists():
            sys.path.append(str(user_site))
    import yaml


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "data" / "macro-indicators.yaml"
OUT_DIR = ROOT / "data" / "macro"
HISTORY_DIR = OUT_DIR / "history"
RAW_DIR = OUT_DIR / "raw"
CRYPTO_OI_DIR = ROOT / "data" / "crypto" / "oi-history"


FRED_SERIES_IDS = {
    "FEDFUNDS",
    "DGS2",
    "DGS10",
    "DFII10",
    "T10Y2Y",
    "GDPC1",
    "NAPM",
    "RSAFS",
    "CPIAUCSL",
    "CPILFESL",
    "PCEPILFE",
    "DCOILWTICO",
    "UNRATE",
    "PAYEMS",
    "ICSA",
    "M2SL",
    "WALCL",
    "WRESBAL",
    "BAMLH0A0HYM2",
    "MORTGAGE30US",
    "CSUSHPINSA",
    "DEXUSEU",
    "DEXJPUS",
    "DCOILBRENTEU",
    "GOLDAMGBD228NLBM",
    "VIXCLS",
    "SP500",
    "DEXCHUS",
    "ECBDFR",
    "IRLTLT01JPM156N",
    "DTWEXBGS",
    "CPALTT01CNM659N",
    "PCOPPUSDM",
    "CPATAX",
    "BSCICP03EZM665S",
    "DEUPROINDMISMEI",
}


STOOQ_SYMBOLS = {
    "us_dxy": "dx.f",
    "global_copper_price": "hg.f",
}


EUROSTAT_SERIES = {
    "eu_core_hicp": {
        "dataset": "prc_hicp_manr",
        "params": {
            "lang": "en",
            "freq": "M",
            "unit": "RCH_A",
            "coicop": "TOT_X_NRG_FOOD",
            "geo": "EA20",
        },
        "source_name": "Eurostat prc_hicp_manr / HICP annual rate of change",
        "gap": "Official Eurostat core HICP YoY: overall index excluding energy, food, alcohol and tobacco for euro area 20.",
    },
}


COINGECKO_IDS = {
    "xasset_btc_price_trend": "bitcoin",
    "xasset_eth_btc": "ethereum",
}


NASDAQ_SYMBOLS = {
    "xasset_gold_price": ("GLD", "etf"),
    "xasset_nasdaq100": ("QQQ", "etf"),
    "xasset_msci_acwi": ("ACWI", "etf"),
    "xasset_hyg": ("HYG", "etf"),
    "xasset_reit_index": ("VNQ", "etf"),
}


TREASURY_YIELD_INDICATORS = {
    "us_10y_treasury_yield": ("10 Yr", "DGS10"),
    "us_2y10y_spread": ("10y_minus_2y", "T10Y2Y"),
}


TREASURY_REAL_YIELD_INDICATORS = {
    "us_10y_real_yield": ("10 YR", "DFII10"),
}


HYPERLIQUID_COINS = ["BTC", "ETH", "SOL"]


MACROPAGE_XLSX_SERIES = {
    "cn_tsf_stock_yoy": {
        "file": "macropage_credit_money.xlsx",
        "column": "社会融资规模存量",
        "operation": "yoy",
        "source_name": "AFAN-LIFE/macropage 社融货币.xlsx / PBOC",
    },
    "cn_tsf_flow": {
        "file": "macropage_credit_money.xlsx",
        "column": "社会融资增量:当月值",
        "source_name": "AFAN-LIFE/macropage 社融货币.xlsx / PBOC",
    },
    "cn_m1_yoy": {
        "file": "macropage_credit_money.xlsx",
        "column": "M1:同比",
        "source_name": "AFAN-LIFE/macropage 社融货币.xlsx / PBOC",
    },
    "cn_m2_yoy": {
        "file": "macropage_credit_money.xlsx",
        "column": "M2:同比",
        "source_name": "AFAN-LIFE/macropage 社融货币.xlsx / PBOC",
    },
    "cn_new_rmb_loans": {
        "file": "macropage_credit_money.xlsx",
        "column": "社会融资增量:新增人民币贷款:当月值",
        "source_name": "AFAN-LIFE/macropage 社融货币.xlsx / PBOC",
    },
    "cn_corporate_medium_long_loans": {
        "file": "macropage_credit_money.xlsx",
        "column": "金融机构:新增人民币贷款:非金融性公司及其他部门:中长期:当月值",
        "source_name": "AFAN-LIFE/macropage 社融货币.xlsx / PBOC",
    },
    "cn_household_medium_long_loans": {
        "file": "macropage_credit_money.xlsx",
        "column": "金融机构:新增人民币贷款:居民户:中长期:当月值",
        "source_name": "AFAN-LIFE/macropage 社融货币.xlsx / PBOC",
    },
    "cn_official_manufacturing_pmi": {
        "file": "macropage_pmi.xlsx",
        "column": "PMI",
        "source_name": "AFAN-LIFE/macropage PMI.xlsx / NBS",
    },
    "cn_fixed_asset_investment_yoy": {
        "file": "macropage_investment.xlsx",
        "column": "固定资产投资完成额:累计同比",
        "source_name": "AFAN-LIFE/macropage 固定资产投资.xlsx / NBS",
    },
    "cn_real_estate_development_investment_yoy": {
        "file": "macropage_real_estate.xlsx",
        "column": "房地产开发投资完成额:累计同比",
        "source_name": "AFAN-LIFE/macropage 房地产投资.xlsx / NBS",
    },
    "cn_commercial_housing_sales_area_yoy": {
        "file": "macropage_real_estate.xlsx",
        "column": "商品房销售面积:累计值",
        "operation": "yoy",
        "source_name": "AFAN-LIFE/macropage 房地产投资.xlsx / NBS",
    },
    "cn_70_city_home_price": {
        "file": "macropage_70city.xlsx",
        "operation": "average_columns_contains",
        "contains": "新建商品住宅价格指数",
        "source_name": "AFAN-LIFE/macropage 70城.xlsx / NBS",
    },
    "cn_ppi_yoy": {
        "file": "macropage_cpi_ppi.xlsx",
        "column": "PPI:全部工业品:环比",
        "operation": "mom_chain_yoy",
        "source_name": "AFAN-LIFE/macropage CPI+PPI.xlsx / NBS",
    },
    "cn_retail_sales_yoy": {
        "file": "macropage_retail.xlsx",
        "column": "社会消费品零售总额:当月值",
        "operation": "yoy",
        "source_name": "AFAN-LIFE/macropage 社零.xlsx / NBS",
    },
    "cn_exports_yoy": {
        "file": "macropage_trade.xlsx",
        "column": "出口金额:当月同比",
        "source_name": "AFAN-LIFE/macropage 进出口基本.xlsx / China Customs",
    },
}


MACROPAGE_CSV_SERIES = {
    "cn_10y_gov_bond_yield": {
        "file": "macropage_yield.csv",
        "date_column": "workTime",
        "value_column": "tenYear",
        "source_name": "AFAN-LIFE/macropage yield.csv / ChinaBond yield curve",
    },
}


EASTMONEY_SERIES = {
    "cn_m1_yoy": {
        "report_name": "RPT_ECONOMY_CURRENCY_SUPPLY",
        "value_field": "CURRENCY_SAME",
        "source_name": "Eastmoney Datacenter / PBOC money supply",
        "gap": "Public Eastmoney mirror of PBOC money supply; uses M1 YoY field.",
    },
    "cn_m2_yoy": {
        "report_name": "RPT_ECONOMY_CURRENCY_SUPPLY",
        "value_field": "BASIC_CURRENCY_SAME",
        "source_name": "Eastmoney Datacenter / PBOC money supply",
        "gap": "Public Eastmoney mirror of PBOC money supply; uses M2 YoY field.",
    },
    "cn_official_manufacturing_pmi": {
        "report_name": "RPT_ECONOMY_PMI",
        "value_field": "MAKE_INDEX",
        "source_name": "Eastmoney Datacenter / NBS-CFLP PMI",
        "gap": "Public Eastmoney mirror of official manufacturing PMI.",
    },
    "cn_cpi_yoy": {
        "report_name": "RPT_ECONOMY_CPI",
        "value_field": "NATIONAL_SAME",
        "source_name": "Eastmoney Datacenter / NBS CPI",
        "gap": "Public Eastmoney mirror of NBS CPI YoY.",
    },
    "cn_ppi_yoy": {
        "report_name": "RPT_ECONOMY_PPI",
        "value_field": "BASE_SAME",
        "source_name": "Eastmoney Datacenter / NBS PPI",
        "gap": "Public Eastmoney mirror of NBS PPI YoY; replaces previous derived chain YoY proxy.",
    },
    "cn_retail_sales_yoy": {
        "report_name": "RPT_ECONOMY_TOTAL_RETAIL",
        "value_field": "RETAIL_TOTAL_SAME",
        "source_name": "Eastmoney Datacenter / NBS total retail sales",
        "gap": "Public Eastmoney mirror of NBS monthly total retail sales YoY.",
    },
    "cn_exports_yoy": {
        "report_name": "RPT_ECONOMY_CUSTOMS",
        "value_field": "EXIT_BASE_SAME",
        "source_name": "Eastmoney Datacenter / China Customs",
        "gap": "Public Eastmoney mirror of China Customs export YoY.",
    },
    "cn_new_rmb_loans": {
        "report_name": "RPT_ECONOMY_RMB_LOAN",
        "value_field": "RMB_LOAN",
        "source_name": "Eastmoney Datacenter / PBOC new RMB loans",
        "gap": "Public Eastmoney mirror of monthly new RMB loans.",
    },
    "cn_industrial_production_yoy": {
        "report_name": "RPT_ECONOMY_INDUS_GROW",
        "value_field": "BASE_SAME",
        "source_name": "Eastmoney Datacenter / NBS industrial production",
        "gap": "Public Eastmoney mirror of NBS value-added industrial output YoY.",
    },
    "cn_fixed_asset_investment_yoy": {
        "report_name": "RPT_ECONOMY_ASSET_INVEST",
        "value_field": "BASE_SAME",
        "source_name": "Eastmoney Datacenter / NBS fixed asset investment",
        "gap": "Public Eastmoney mirror of fixed-asset investment YoY; field is monthly YoY, not cumulative YoY.",
    },
}


EASTMONEY_HOUSE_PRICE_SERIES = {
    "cn_70_city_home_price": {
        "report_name": "RPT_ECONOMY_HOUSE_PRICE",
        "value_field": "FIRST_COMHOUSE_SAME",
        "source_name": "Eastmoney Datacenter / NBS 70-city home price",
        "gap": "Public Eastmoney mirror of NBS 70-city new commercial housing price YoY index; value is average across available cities for each month.",
    },
}


MOFCOM_SOCIAL_FINANCING_FLOW = {
    "cn_tsf_flow": {
        "url": "https://data.mofcom.gov.cn/datamofcom/front/gnmy/shrzgmQuery",
        "value_field": "tiosfs",
        "source_name": "MOFCOM data mirror / PBOC aggregate financing flow",
        "gap": "MOFCOM domestic-trade data mirror of social financing flow; latest may lag PBOC by one release month.",
    },
}


PBOC_STATS_SERIES = {
    "cn_tsf_stock_yoy": {
        "index_url": "https://www.pbc.gov.cn/diaochatongjisi/116219/116319/2026ntjsj/shrzgm/index.html",
        "table_title": "社会融资规模存量统计表",
        "row_label": "社会融资规模存量",
        "value_kind": "growth",
        "seed": "macropage",
        "source_name": "PBOC official statistics / AFRE stock",
        "gap": "Official PBOC 2026 statistics table patched onto macropage long-history cache; PBOC files are annual workbooks with latest months filled in.",
    },
    "cn_household_medium_long_loans": {
        "index_url": "https://www.pbc.gov.cn/diaochatongjisi/116219/116319/2026ntjsj/jrjgxdsztj/index.html",
        "table_title": "金融机构本外币信贷收支表",
        "parent_label": "住户贷款",
        "row_label": "中长期贷款",
        "value_kind": "balance_monthly_change",
        "seed": "macropage",
        "source_name": "PBOC official statistics / Household medium and long-term loan balance change",
        "gap": "Official PBOC credit-balance table patched onto macropage flow history. Latest points are monthly balance changes, a proxy for monthly new medium/long-term loans and may differ from strict flow statistics.",
    },
    "cn_corporate_medium_long_loans": {
        "index_url": "https://www.pbc.gov.cn/diaochatongjisi/116219/116319/2026ntjsj/jrjgxdsztj/index.html",
        "table_title": "金融机构本外币信贷收支表",
        "parent_label": "企（事）业单位贷款",
        "row_label": "中长期贷款",
        "value_kind": "balance_monthly_change",
        "seed": "macropage",
        "source_name": "PBOC official statistics / Corporate medium and long-term loan balance change",
        "gap": "Official PBOC credit-balance table patched onto macropage flow history. Latest points are monthly balance changes, a proxy for monthly new medium/long-term loans and may differ from strict flow statistics.",
    },
}


NBS_RELEASE_INDEX_URL = "https://www.stats.gov.cn/sj/zxfb/"


NBS_RELEASE_SERIES = {
    "cn_industrial_production_yoy": {
        "title_keywords": ["规模以上工业增加值"],
        "seed": "eastmoney",
        "parse": "xlsx_row",
        "row_label": "规模以上工业增加值",
        "value_column": 3,
        "source_name": "NBS official data release / Industrial value-added output",
        "gap": "Official NBS latest release patched onto Eastmoney long-history mirror; NBS structured easyquery endpoint is WAF-blocked in this environment.",
    },
    "cn_fixed_asset_investment_yoy": {
        "title_keywords": ["固定资产投资"],
        "seed": "eastmoney",
        "parse": "xlsx_row",
        "row_label": "固定资产投资（不含农户）",
        "value_column": 1,
        "source_name": "NBS official data release / Fixed asset investment",
        "gap": "Official NBS release uses cumulative YoY. Historical seed may contain mirror/vendor revisions.",
    },
    "cn_real_estate_development_investment_yoy": {
        "title_keywords": ["房地产市场基本情况"],
        "seed": "macropage",
        "parse": "xlsx_row",
        "sheet_index": 0,
        "row_label": "房地产开发投资（亿元）",
        "value_column": 2,
        "source_name": "NBS official data release / Real estate development investment",
        "gap": "Official NBS latest release patched onto macropage long-history cache until a full official history path is added.",
    },
    "cn_commercial_housing_sales_area_yoy": {
        "title_keywords": ["房地产市场基本情况"],
        "seed": "macropage",
        "parse": "xlsx_row",
        "sheet_index": 2,
        "row_label": "全国总计",
        "value_column": 2,
        "source_name": "NBS official data release / Commercial housing sales area",
        "gap": "Official NBS latest release patched onto macropage long-history cache until a full official history path is added.",
    },
    "cn_retail_sales_yoy": {
        "title_keywords": ["社会消费品零售总额"],
        "seed": "eastmoney",
        "parse": "retail_text",
        "source_name": "NBS official data release / Total retail sales of consumer goods",
        "gap": "Official NBS monthly YoY parsed from release text; old .xls attachment is not parsed without xlrd.",
    },
    "cn_cpi_yoy": {
        "title_keywords": ["居民消费价格"],
        "seed": "eastmoney",
        "parse": "xlsx_row",
        "row_label": "居民消费价格",
        "value_column": 2,
        "source_name": "NBS official data release / CPI",
        "gap": "Official NBS latest release patched onto Eastmoney long-history mirror.",
    },
    "cn_ppi_yoy": {
        "title_keywords": ["工业生产者出厂价格"],
        "seed": "eastmoney",
        "parse": "xlsx_row",
        "row_label": "一、工业生产者出厂价格",
        "value_column": 2,
        "source_name": "NBS official data release / PPI",
        "gap": "Official NBS latest release patched onto Eastmoney long-history mirror.",
    },
    "cn_official_manufacturing_pmi": {
        "title_keywords": ["采购经理指数运行情况"],
        "seed": "eastmoney",
        "parse": "pmi_text",
        "source_name": "NBS official data release / China manufacturing PMI",
        "gap": "Official NBS manufacturing PMI parsed from release text; old .xls attachment is not parsed without xlrd.",
    },
    "cn_70_city_home_price": {
        "title_keywords": ["70个大中城市商品住宅销售价格变动情况"],
        "seed": "eastmoney_house",
        "parse": "home_price_html_average",
        "source_name": "NBS official data release / 70-city new home price index",
        "gap": "Official NBS latest release table average patched onto Eastmoney long-history mirror.",
    },
}


JIN10_REPORTS = {
    "us_ism_manufacturing_pmi": {
        "attr_id": "28",
        "symbol": "美国ISM制造业PMI报告",
        "all_js_file": "jin10_usa_ism_pmi_all.js",
        "source_url": "https://datacenter.jin10.com/reportType/dc_usa_ism_pmi",
        "official_recent_url": "https://www.ismworld.org/supply-management-news-and-reports/reports/ism-pmi-reports/pmi/may/",
        "official_recent_cache": "ism_manufacturing_pmi_may_2026.html",
        "official_recent_start": "THE LAST 12 MONTHS",
        "official_recent_end": "Average for 12 months",
        "official_recent_source": "ISM official Manufacturing PMI Report, May 2026",
    },
    "us_ism_services_pmi": {
        "attr_id": "29",
        "symbol": "美国ISM非制造业PMI报告",
        "all_js_file": "jin10_usa_ism_non_pmi_all.js",
        "source_url": "https://datacenter.jin10.com/reportType/dc_usa_ism_non_pmi",
        "official_recent_url": "https://www.ismworld.org/supply-management-news-and-reports/reports/ism-pmi-reports/services/may/",
        "official_recent_cache": "ism_services_pmi_may_2026.html",
        "official_recent_start": '<h3 class="text-strong text-uppercase" style="width: 100%;">Services PMI',
        "official_recent_end": "Average for",
        "official_recent_source": "ISM official Services PMI Report, May 2026",
    },
    "cn_caixin_manufacturing_pmi": {
        "attr_id": "73",
        "symbol": "中国财新制造业PMI终值报告",
        "all_js_file": "jin10_chinese_caixin_manufacturing_pmi_all.js",
        "source_url": "https://datacenter.jin10.com/reportType/dc_chinese_caixin_manufacturing_pmi",
    },
    "cn_industrial_production_yoy": {
        "attr_id": "58",
        "symbol": "中国规模以上工业增加值年率报告",
        "all_js_file": "jin10_china_industrial_production_yoy_all.js",
        "source_url": "https://datacenter.jin10.com/reportType/dc_chinese_industrial_production_yoy",
    },
}

MONTH_ABBR_TO_NUMBER = {
    "Jan": "01",
    "Feb": "02",
    "Mar": "03",
    "Apr": "04",
    "May": "05",
    "Jun": "06",
    "Jul": "07",
    "Aug": "08",
    "Sep": "09",
    "Oct": "10",
    "Nov": "11",
    "Dec": "12",
}


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def run_curl(url: str, *, method: str = "GET", data: Any = None, timeout: int = 30) -> str:
    args = ["curl", "-sS", "-L", "--max-time", str(timeout)]
    if method == "POST":
        args.extend(["-H", "Content-Type: application/json", "-d", json.dumps(data)])
    args.append(url)
    completed = subprocess.run(
        args,
        cwd=ROOT,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    if completed.returncode != 0:
        raise RuntimeError(completed.stderr.strip() or f"curl exited {completed.returncode}")
    return completed.stdout


def run_curl_bytes(url: str, *, timeout: int = 30) -> bytes:
    args = ["curl", "-sS", "-L", "--compressed", "--max-time", str(timeout), "-A", "Mozilla/5.0", url]
    completed = subprocess.run(
        args,
        cwd=ROOT,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    if completed.returncode != 0:
        raise RuntimeError(completed.stderr.decode(errors="ignore").strip() or f"curl exited {completed.returncode}")
    return completed.stdout


def fetch_text_live_first(
    url: str,
    cache_path: Path,
    *,
    timeout: int = 45,
    curl_args: list[str] | None = None,
    binary: bool = False,
    reject_html: bool = False,
) -> tuple[str | bytes, dict[str, Any]]:
    try:
        if curl_args is None:
            content = run_curl_bytes(url, timeout=timeout) if binary else run_curl(url, timeout=timeout)
        else:
            completed = subprocess.run(
                curl_args,
                cwd=ROOT,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=not binary,
                check=False,
            )
            if completed.returncode != 0:
                stderr = completed.stderr if isinstance(completed.stderr, str) else completed.stderr.decode(errors="ignore")
                raise RuntimeError(stderr.strip() or f"curl exited {completed.returncode}")
            content = completed.stdout
        if reject_html and not binary:
            text = content if isinstance(content, str) else content.decode(errors="ignore")
            if text.lstrip().startswith("<"):
                raise RuntimeError("live response looked like HTML/error page")
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        if binary:
            cache_path.write_bytes(content if isinstance(content, bytes) else content.encode())
        else:
            cache_path.write_text(content if isinstance(content, str) else content.decode(errors="ignore"))
        return content, {"url": url, "cache_path": str(cache_path.relative_to(ROOT)), "refreshed": True}
    except Exception as exc:  # noqa: BLE001
        if not cache_path.exists():
            raise
        cached = cache_path.read_bytes() if binary else cache_path.read_text(errors="ignore")
        return cached, {
            "url": url,
            "cache_path": str(cache_path.relative_to(ROOT)),
            "refreshed": False,
            "cache_fallback": True,
            "live_error": str(exc),
        }


def parse_float(value: str | None) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text or text == "." or text.lower() in {"nan", "null", "none"}:
        return None
    try:
        value_float = float(text)
    except ValueError:
        return None
    if not math.isfinite(value_float):
        return None
    return value_float


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").replace("\u3000", " ").replace("\xa0", " ")).strip()


def compact_history(rows: list[dict[str, Any]], max_points: int | None) -> list[dict[str, Any]]:
    if not max_points or len(rows) <= max_points:
        return rows
    return rows[-max_points:]


def summarize_history(rows: list[dict[str, Any]]) -> dict[str, Any]:
    valid = [row for row in rows if row.get("value") is not None]
    if not valid:
        return {
            "status": "missing",
            "points": 0,
            "first_date": None,
            "latest_date": None,
            "latest_value": None,
        }
    latest = valid[-1]
    return {
        "status": "ok",
        "points": len(valid),
        "first_date": valid[0]["date"],
        "latest_date": latest["date"],
        "latest_value": latest["value"],
    }


def fetch_fred(series_id: str, *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?{urlencode({'id': series_id})}"
    cache_path = RAW_DIR / f"{series_id}.csv"
    text, meta = fetch_text_live_first(url, cache_path, timeout=45, reject_html=True)
    assert isinstance(text, str)
    source_meta = {"adapter": "fred_csv" if meta.get("refreshed") else "fred_csv_cache_fallback", **meta}
    if text.lstrip().startswith("<"):
        raise RuntimeError("FRED returned HTML instead of CSV")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames or len(reader.fieldnames) < 2:
        raise RuntimeError("FRED CSV has no data columns")
    date_key = reader.fieldnames[0]
    value_key = reader.fieldnames[1]
    rows = []
    for row in reader:
        value = parse_float(row.get(value_key))
        if value is None:
            continue
        rows.append({"date": row[date_key], "value": value})
    return compact_history(rows, max_points), source_meta


def fetch_stooq(symbol: str, *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    url = f"https://stooq.com/q/d/l/?{urlencode({'s': symbol, 'i': 'd'})}"
    cache_path = RAW_DIR / f"{symbol}.csv"
    text, meta = fetch_text_live_first(url, cache_path, timeout=45, reject_html=True)
    assert isinstance(text, str)
    source_meta = {"adapter": "stooq_csv" if meta.get("refreshed") else "stooq_csv_cache_fallback", "symbol": symbol, **meta}
    if not text.strip() or "No data" in text:
        raise RuntimeError("Stooq returned no data")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        value = parse_float(row.get("Close"))
        date = row.get("Date")
        if date and value is not None:
            rows.append({"date": date, "value": value})
    return compact_history(rows, max_points), source_meta


def fetch_eurostat_jsonstat(spec: dict[str, Any], *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    query = urlencode(spec["params"])
    url = f"https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/{spec['dataset']}?{query}"
    cache_path = RAW_DIR / f"eurostat_{spec['dataset']}_{spec['params']['geo']}_{spec['params']['coicop']}.json"
    text, meta = fetch_text_live_first(url, cache_path, timeout=60, reject_html=True)
    assert isinstance(text, str)
    payload = json.loads(text)
    source_meta = {"adapter": "eurostat_jsonstat" if meta.get("refreshed") else "eurostat_jsonstat_cache_fallback", **meta}

    if "error" in payload:
        raise RuntimeError(f"Eurostat returned error: {payload['error']}")
    dimension = payload.get("dimension") or {}
    time_category = ((dimension.get("time") or {}).get("category") or {})
    time_index = time_category.get("index") or {}
    time_labels = time_category.get("label") or {}
    values = payload.get("value") or {}
    if not time_index:
        raise RuntimeError("Eurostat JSON-stat payload has no time index")

    rows = []
    for period, index in sorted(time_index.items(), key=lambda item: item[1]):
        value = parse_float(values.get(str(index)))
        if value is None:
            continue
        date = f"{period}-01" if len(period) == 7 else time_labels.get(period, period)
        rows.append({"date": date, "value": value})

    return compact_history(rows, max_points), {
        **source_meta,
        "source_name": spec.get("source_name"),
        "dataset": spec["dataset"],
        "params": spec["params"],
        "updated": payload.get("updated"),
        "label": payload.get("label"),
        "gap": spec.get("gap"),
    }


def eastmoney_report_cache_path(report_name: str) -> Path:
    return RAW_DIR / f"eastmoney_{report_name}.json"


def fetch_eastmoney_report(report_name: str, *, page_size: int = 5000) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    base_url = "https://datacenter-web.eastmoney.com/api/data/v1/get"
    cache_path = eastmoney_report_cache_path(report_name)
    source_meta: dict[str, Any]
    try:
        first_url = f"{base_url}?{urlencode({'reportName': report_name, 'columns': 'ALL', 'pageNumber': 1, 'pageSize': page_size, 'sortColumns': 'REPORT_DATE', 'sortTypes': -1, 'source': 'WEB', 'client': 'WEB'})}"
        first_payload = json.loads(run_curl(first_url, timeout=30))
        result = first_payload.get("result") or {}
        pages = int(result.get("pages") or 1)
        data = list(result.get("data") or [])
        for page_number in range(2, pages + 1):
            url = f"{base_url}?{urlencode({'reportName': report_name, 'columns': 'ALL', 'pageNumber': page_number, 'pageSize': page_size, 'sortColumns': 'REPORT_DATE', 'sortTypes': -1, 'source': 'WEB', 'client': 'WEB'})}"
            payload = json.loads(run_curl(url, timeout=30))
            data.extend(((payload.get("result") or {}).get("data") or []))
        cache_path.write_text(json.dumps({"data": data}, ensure_ascii=False, indent=2) + "\n")
        source_meta = {
            "adapter": "eastmoney_public_api",
            "report_name": report_name,
            "url": first_url,
            "cache_path": str(cache_path.relative_to(ROOT)),
            "pages": pages,
        }
    except Exception as exc:  # noqa: BLE001
        if not cache_path.exists():
            raise
        data = json.loads(cache_path.read_text()).get("data") or []
        source_meta = {
            "adapter": "eastmoney_public_api_cache",
            "report_name": report_name,
            "cache_path": str(cache_path.relative_to(ROOT)),
            "fallback_error": str(exc),
        }
    return data, source_meta


def fetch_eastmoney_series(spec: dict[str, Any], *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    data, meta = fetch_eastmoney_report(spec["report_name"])
    rows_by_date: dict[str, float] = {}
    for item in data:
        raw_date = item.get("REPORT_DATE")
        value = parse_float(item.get(spec["value_field"]))
        if not raw_date or value is None:
            continue
        rows_by_date[str(raw_date)[:10]] = value
    rows = [{"date": date, "value": value} for date, value in sorted(rows_by_date.items())]
    return compact_history(rows, max_points), {
        **meta,
        "source_name": spec.get("source_name"),
        "value_field": spec["value_field"],
        "gap": spec.get("gap"),
    }


def fetch_eastmoney_house_price_average(spec: dict[str, Any], *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    data, meta = fetch_eastmoney_report(spec["report_name"])
    values_by_date: dict[str, list[float]] = {}
    for item in data:
        raw_date = item.get("REPORT_DATE")
        value = parse_float(item.get(spec["value_field"]))
        if not raw_date or value is None:
            continue
        values_by_date.setdefault(str(raw_date)[:10], []).append(value)
    rows = [
        {"date": date, "value": sum(values) / len(values)}
        for date, values in sorted(values_by_date.items())
        if values
    ]
    return compact_history(rows, max_points), {
        **meta,
        "source_name": spec.get("source_name"),
        "value_field": spec["value_field"],
        "operation": "average_by_month",
        "gap": spec.get("gap"),
    }


def fetch_mofcom_social_financing_flow(spec: dict[str, Any], *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    cache_path = RAW_DIR / "mofcom_shrzgm.json"
    try:
        args = ["curl", "-sS", "-L", "--max-time", "30", "-X", "POST", spec["url"]]
        completed = subprocess.run(
            args,
            cwd=ROOT,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )
        if completed.returncode != 0:
            raise RuntimeError(completed.stderr.strip() or f"curl exited {completed.returncode}")
        payload = json.loads(completed.stdout)
        cache_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n")
        source_meta = {
            "adapter": "mofcom_social_financing_flow_api",
            "url": spec["url"],
            "cache_path": str(cache_path.relative_to(ROOT)),
        }
    except Exception as exc:  # noqa: BLE001
        if not cache_path.exists():
            raise
        payload = json.loads(cache_path.read_text())
        source_meta = {
            "adapter": "mofcom_social_financing_flow_cache",
            "url": spec["url"],
            "cache_path": str(cache_path.relative_to(ROOT)),
            "fallback_error": str(exc),
        }
    rows = []
    for item in payload or []:
        period = str(item.get("date") or "")
        value = parse_float(item.get(spec["value_field"]))
        if len(period) == 6 and value is not None:
            rows.append({"date": f"{period[:4]}-{period[4:]}-01", "value": value})
    rows.sort(key=lambda item: item["date"])
    return compact_history(rows, max_points), {
        **source_meta,
        "source_name": spec.get("source_name"),
        "value_field": spec["value_field"],
        "gap": spec.get("gap"),
    }


def fetch_pboc_stats_page(url: str, indicator_id: str) -> tuple[str, dict[str, Any]]:
    cache_path = RAW_DIR / f"pbc_{indicator_id}_index.html"
    try:
        text = run_curl(url, timeout=45)
        cache_path.write_text(text, encoding="utf-8")
        source_meta = {
            "adapter": "pboc_stats_index",
            "url": url,
            "cache_path": str(cache_path.relative_to(ROOT)),
        }
    except Exception as exc:  # noqa: BLE001
        if not cache_path.exists():
            raise
        text = cache_path.read_text(encoding="utf-8", errors="ignore")
        source_meta = {
            "adapter": "pboc_stats_index_cache",
            "url": url,
            "cache_path": str(cache_path.relative_to(ROOT)),
            "fallback_error": str(exc),
        }
    return text, source_meta


def pboc_table_xlsx_url(index_url: str, html_text: str, table_title: str) -> str:
    rows = re.findall(r"<tr[\s\S]*?</tr>", html_text, flags=re.IGNORECASE)
    for row in rows:
        plain = normalize_text(re.sub(r"<[^>]+>", " ", row))
        if normalize_text(table_title) not in plain:
            continue
        match = re.search(r'href="([^"]+\.xlsx)"', row, flags=re.IGNORECASE)
        if match:
            return urljoin(index_url, match.group(1))
    raise RuntimeError(f"PBOC xlsx link not found for table: {table_title}")


def download_pboc_xlsx(url: str, indicator_id: str) -> tuple[Path, dict[str, Any]]:
    stamp_match = re.search(r"/(\d{6})\d+\.xlsx$", url)
    stamp = stamp_match.group(1) if stamp_match else datetime.now(timezone.utc).strftime("%Y%m")
    cache_path = RAW_DIR / f"pbc_{indicator_id}_{stamp}.xlsx"
    if cache_path.exists():
        return cache_path, {
            "adapter": "pboc_stats_xlsx_cache",
            "url": url,
            "cache_path": str(cache_path.relative_to(ROOT)),
        }
    cache_path.write_bytes(run_curl_bytes(url, timeout=45))
    return cache_path, {
        "adapter": "pboc_stats_xlsx",
        "url": url,
        "cache_path": str(cache_path.relative_to(ROOT)),
    }


def parse_pboc_afre_stock_growth(path: Path, row_label: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for PBOC xlsx parsing") from exc

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    month_row = next((row for row in rows if any(re.fullmatch(r"20\d{2}\.\d{1,2}", str(value or "")) for value in row)), None)
    if month_row is None:
        raise RuntimeError("PBOC AFRE stock workbook has no month header row")
    target = normalize_text(row_label)
    data_row = None
    for row in rows:
        if row and normalize_text(row[0]) == target:
            data_row = row
            break
    if data_row is None:
        raise RuntimeError(f"PBOC AFRE stock row not found: {row_label}")

    parsed_rows = []
    for index, month_value in enumerate(month_row):
        month_match = re.fullmatch(r"(20\d{2})\.(\d{1,2})", str(month_value or ""))
        if not month_match:
            continue
        year = int(month_match.group(1))
        month = int(month_match.group(2))
        if not 1 <= month <= 12:
            continue
        growth_index = index + 1
        value = parse_float(data_row[growth_index] if growth_index < len(data_row) else None)
        if value is not None:
            parsed_rows.append({"date": f"{year:04d}-{month:02d}-01", "value": value})
    return parsed_rows


def pboc_month_headers(row: tuple[Any, ...]) -> list[tuple[int, str]]:
    headers = []
    for index, value in enumerate(row):
        month_match = re.fullmatch(r"(20\d{2})\.(\d{1,2})", str(value or ""))
        if not month_match:
            continue
        year = int(month_match.group(1))
        month = int(month_match.group(2))
        if 1 <= month <= 12:
            headers.append((index, f"{year:04d}-{month:02d}-01"))
    return headers


def parse_pboc_credit_balance_change(path: Path, parent_label: str, row_label: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for PBOC xlsx parsing") from exc

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    month_row = next((row for row in rows if len(pboc_month_headers(row)) >= 2), None)
    if month_row is None:
        raise RuntimeError("PBOC credit workbook has no month header row")
    headers = pboc_month_headers(month_row)

    normalized_parent = normalize_text(parent_label)
    normalized_row = normalize_text(row_label)
    in_parent_block = False
    target_row = None
    for row in rows:
        label = normalize_text(row[0] if row else None)
        if normalized_parent in label:
            in_parent_block = True
            continue
        if in_parent_block and normalized_row in label:
            target_row = row
            break
    if target_row is None:
        raise RuntimeError(f"PBOC credit row not found: {parent_label} / {row_label}")

    values = [(date, parse_float(target_row[index] if index < len(target_row) else None)) for index, date in headers]
    rows_out = []
    previous = None
    for date, value in values:
        if value is not None and previous is not None:
            rows_out.append({"date": date, "value": round(value - previous, 6)})
        if value is not None:
            previous = value
    return rows_out


def fetch_pboc_stats_series(indicator_id: str, spec: dict[str, Any], *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if spec.get("seed") == "macropage":
        seed_rows, seed_meta = fetch_macropage_xlsx(MACROPAGE_XLSX_SERIES[indicator_id], max_points=None)
    else:
        raise RuntimeError(f"Unsupported PBOC history seed: {spec.get('seed')}")
    index_html, index_meta = fetch_pboc_stats_page(spec["index_url"], indicator_id)
    xlsx_url = pboc_table_xlsx_url(spec["index_url"], index_html, spec["table_title"])
    xlsx_path, xlsx_meta = download_pboc_xlsx(xlsx_url, indicator_id)
    if spec.get("value_kind") == "growth":
        official_rows = parse_pboc_afre_stock_growth(xlsx_path, spec["row_label"])
    elif spec.get("value_kind") == "balance_monthly_change":
        official_rows = parse_pboc_credit_balance_change(xlsx_path, spec["parent_label"], spec["row_label"])
    else:
        raise RuntimeError(f"Unsupported PBOC value kind: {spec.get('value_kind')}")

    rows_by_date = {row["date"]: row["value"] for row in seed_rows if row.get("value") is not None}
    rows_by_date.update({row["date"]: row["value"] for row in official_rows if row.get("value") is not None})
    rows = [{"date": date, "value": value} for date, value in sorted(rows_by_date.items())]
    return compact_history(rows, max_points), {
        "adapter": "pboc_official_stats_plus_history_seed",
        "source_name": spec.get("source_name"),
        "table_title": spec["table_title"],
        "components": {
            "history_seed": seed_meta,
            "stats_index": index_meta,
            "stats_xlsx": xlsx_meta,
        },
        "gap": spec.get("gap"),
    }


def fetch_nbs_release_index() -> tuple[list[dict[str, str]], dict[str, Any]]:
    cache_path = RAW_DIR / "nbs_stats_zxfb.html"
    try:
        text = run_curl(NBS_RELEASE_INDEX_URL, timeout=45)
        cache_path.write_text(text, encoding="utf-8")
        source_meta = {
            "adapter": "nbs_release_index",
            "url": NBS_RELEASE_INDEX_URL,
            "cache_path": str(cache_path.relative_to(ROOT)),
        }
    except Exception as exc:  # noqa: BLE001
        if not cache_path.exists():
            raise
        text = cache_path.read_text(encoding="utf-8", errors="ignore")
        source_meta = {
            "adapter": "nbs_release_index_cache",
            "url": NBS_RELEASE_INDEX_URL,
            "cache_path": str(cache_path.relative_to(ROOT)),
            "fallback_error": str(exc),
        }

    links = []
    seen = set()
    for match in re.finditer(r'<a[^>]+href="([^"]+)"[^>]*title=[\'"]([^\'"]+)[\'"]', text):
        href, title = match.groups()
        if not title or "javascript" in href:
            continue
        url = urljoin(NBS_RELEASE_INDEX_URL, href)
        key = (title, url)
        if key in seen:
            continue
        seen.add(key)
        links.append({"title": title, "url": url})
    return links, source_meta


def nbs_release_period(title: str) -> str:
    match = re.search(r"(\d{4})年(?:\d+[—\-－])?(\d{1,2})月", title)
    if not match:
        raise RuntimeError(f"Cannot parse NBS release period from title: {title}")
    year, month = match.groups()
    return f"{int(year):04d}-{int(month):02d}-01"


def find_nbs_release(spec: dict[str, Any]) -> tuple[dict[str, str], dict[str, Any]]:
    links, meta = fetch_nbs_release_index()
    keywords = spec["title_keywords"]
    matches = [
        link for link in links
        if all(keyword in link["title"] for keyword in keywords)
    ]
    if not matches:
        raise RuntimeError(f"No NBS release found for keywords: {keywords}")
    matches.sort(key=lambda item: nbs_release_period(item["title"]), reverse=True)
    return matches[0], meta


def fetch_nbs_release_page(url: str, indicator_id: str, date: str) -> tuple[str, dict[str, Any]]:
    cache_path = RAW_DIR / f"nbs_{indicator_id}_{date[:7].replace('-', '')}.html"
    try:
        text = run_curl(url, timeout=45)
        cache_path.write_text(text, encoding="utf-8")
        source_meta = {
            "adapter": "nbs_release_page",
            "url": url,
            "cache_path": str(cache_path.relative_to(ROOT)),
        }
    except Exception as exc:  # noqa: BLE001
        if not cache_path.exists():
            raise
        text = cache_path.read_text(encoding="utf-8", errors="ignore")
        source_meta = {
            "adapter": "nbs_release_page_cache",
            "url": url,
            "cache_path": str(cache_path.relative_to(ROOT)),
            "fallback_error": str(exc),
        }
    return text, source_meta


def nbs_page_plain_text(html_text: str) -> str:
    text = re.sub(r"<script[\s\S]*?</script>", "", html_text, flags=re.IGNORECASE)
    text = re.sub(r"<style[\s\S]*?</style>", "", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = text.replace("&nbsp;", " ").replace("&mdash;", "—")
    return re.sub(r"\s+", "", text)


def nbs_attachment_url(page_url: str, html_text: str, suffix: str = "xlsx") -> str | None:
    pattern = rf'href="([^"]+\.{re.escape(suffix)})"'
    match = re.search(pattern, html_text, flags=re.IGNORECASE)
    if not match:
        return None
    return urljoin(page_url, match.group(1))


def download_nbs_xlsx(url: str, indicator_id: str, date: str) -> tuple[Path, dict[str, Any]]:
    cache_path = RAW_DIR / f"nbs_{indicator_id}_{date[:7].replace('-', '')}.xlsx"
    if cache_path.exists():
        return cache_path, {
            "adapter": "nbs_release_xlsx_cache",
            "url": url,
            "cache_path": str(cache_path.relative_to(ROOT)),
        }
    cache_path.write_bytes(run_curl_bytes(url, timeout=45))
    return cache_path, {
        "adapter": "nbs_release_xlsx",
        "url": url,
        "cache_path": str(cache_path.relative_to(ROOT)),
    }


def parse_nbs_xlsx_row(path: Path, spec: dict[str, Any]) -> float:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for NBS xlsx release parsing") from exc

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_index = int(spec.get("sheet_index", 0))
    sheet = workbook[workbook.sheetnames[sheet_index]]
    target = normalize_text(spec["row_label"])
    value_column = int(spec["value_column"])
    for raw_row in sheet.iter_rows(values_only=True):
        if not raw_row:
            continue
        labels = {normalize_text(cell) for cell in raw_row}
        if target in labels:
            value = parse_float(raw_row[value_column] if value_column < len(raw_row) else None)
            if value is None:
                raise RuntimeError(f"NBS row {spec['row_label']} has no numeric value in column {value_column}")
            return value
    raise RuntimeError(f"NBS xlsx row not found: {spec['row_label']}")


def parse_signed_percent(text: str, pattern: str) -> float:
    match = re.search(pattern, text)
    if not match:
        raise RuntimeError(f"Cannot parse NBS release text with pattern: {pattern}")
    direction, value_text = match.groups()
    value = parse_float(value_text)
    if value is None:
        raise RuntimeError(f"Cannot parse numeric percentage: {value_text}")
    return -value if direction in {"下降", "减少"} else value


def parse_nbs_home_price_average(html_text: str) -> float:
    try:
        from lxml import html
    except ImportError as exc:
        raise RuntimeError("lxml is required for NBS home-price HTML table parsing") from exc

    doc = html.fromstring(html_text)
    tables = doc.xpath("//table")
    if not tables:
        raise RuntimeError("No NBS home-price tables found")
    values = []
    for row in tables[0].xpath(".//tr")[2:]:
        cells = [normalize_text("".join(cell.xpath(".//text()"))) for cell in row.xpath("./td|./th")]
        # The table is laid out as two repeated groups: city, MoM, YoY, YTD average.
        for offset in (0, 4):
            if len(cells) > offset + 2 and cells[offset] and cells[offset] != "城市":
                value = parse_float(cells[offset + 2])
                if value is not None:
                    values.append(value)
    if not values:
        raise RuntimeError("No numeric NBS home-price YoY index values found")
    return sum(values) / len(values)


def nbs_seed_rows(indicator_id: str, seed: str, *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if seed == "eastmoney":
        return fetch_eastmoney_series(EASTMONEY_SERIES[indicator_id], max_points=max_points)
    if seed == "eastmoney_house":
        return fetch_eastmoney_house_price_average(EASTMONEY_HOUSE_PRICE_SERIES[indicator_id], max_points=max_points)
    if seed == "macropage":
        return fetch_macropage_xlsx(MACROPAGE_XLSX_SERIES[indicator_id], max_points=max_points)
    raise RuntimeError(f"Unsupported NBS seed: {seed}")


def fetch_nbs_release_series(indicator_id: str, spec: dict[str, Any], *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    seed_rows, seed_meta = nbs_seed_rows(indicator_id, spec["seed"], max_points=None)
    release, index_meta = find_nbs_release(spec)
    date = nbs_release_period(release["title"])
    html_text, page_meta = fetch_nbs_release_page(release["url"], indicator_id, date)

    parse_mode = spec["parse"]
    attachment_meta: dict[str, Any] = {}
    if parse_mode == "xlsx_row":
        xlsx_url = nbs_attachment_url(release["url"], html_text, "xlsx")
        if not xlsx_url:
            raise RuntimeError(f"No NBS xlsx attachment found for {release['title']}")
        xlsx_path, attachment_meta = download_nbs_xlsx(xlsx_url, indicator_id, date)
        latest_value = parse_nbs_xlsx_row(xlsx_path, spec)
    elif parse_mode == "retail_text":
        text = nbs_page_plain_text(html_text)
        latest_value = parse_signed_percent(text, r"。5月份，社会消费品零售总额\d+亿元，同比(增长|下降|减少)([0-9.]+)%")
    elif parse_mode == "pmi_text":
        text = nbs_page_plain_text(html_text)
        match = re.search(r"制造业采购经理指数（PMI）为([0-9.]+)%", text)
        if not match:
            raise RuntimeError("Cannot parse official manufacturing PMI from NBS release text")
        latest_value = parse_float(match.group(1))
        if latest_value is None:
            raise RuntimeError("Official manufacturing PMI is not numeric")
    elif parse_mode == "home_price_html_average":
        latest_value = parse_nbs_home_price_average(html_text)
    else:
        raise RuntimeError(f"Unsupported NBS parse mode: {parse_mode}")

    rows_by_date = {row["date"]: row["value"] for row in seed_rows if row.get("value") is not None}
    rows_by_date[date] = latest_value
    rows = [{"date": row_date, "value": value} for row_date, value in sorted(rows_by_date.items())]
    return compact_history(rows, max_points), {
        "adapter": "nbs_official_release_plus_history_seed",
        "source_name": spec.get("source_name"),
        "release_title": release["title"],
        "release_url": release["url"],
        "release_date": date,
        "latest_value_source": "official_nbs_release",
        "components": {
            "history_seed": seed_meta,
            "release_index": index_meta,
            "release_page": page_meta,
            "attachment": attachment_meta or None,
        },
        "parse_mode": parse_mode,
        "gap": spec.get("gap"),
    }


def fetch_nasdaq_history(symbol: str, assetclass: str, *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    url = (
        f"https://api.nasdaq.com/api/quote/{symbol}/historical?"
        f"{urlencode({'assetclass': assetclass, 'fromdate': '1900-01-01', 'todate': datetime.now(timezone.utc).date().isoformat(), 'limit': '9999'})}"
    )
    cache_path = RAW_DIR / f"nasdaq_{symbol}.json"
    args = [
        "curl",
        "-sS",
        "-L",
        "--max-time",
        "45",
        "-H",
        "user-agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/125 Safari/537.36",
        "-H",
        "accept: application/json,text/plain,*/*",
        "-H",
        "origin: https://www.nasdaq.com",
        "-H",
        "referer: https://www.nasdaq.com/",
        url,
    ]
    text, meta = fetch_text_live_first(url, cache_path, timeout=45, curl_args=args, reject_html=True)
    assert isinstance(text, str)
    source_meta = {"adapter": "nasdaq_history" if meta.get("refreshed") else "nasdaq_history_cache_fallback", "symbol": symbol, **meta}
    payload = json.loads(text)
    raw_rows = (((payload.get("data") or {}).get("tradesTable") or {}).get("rows") or [])
    rows = []
    for row in raw_rows:
        date_text = row.get("date")
        value = parse_float(row.get("close"))
        if not date_text or value is None:
            continue
        date = datetime.strptime(date_text, "%m/%d/%Y").date().isoformat()
        rows.append({"date": date, "value": value})
    rows.sort(key=lambda item: item["date"])
    return compact_history(rows, max_points), source_meta


def fetch_treasury_yield_curve(field: str, *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows = []
    files = sorted(RAW_DIR.glob("treasury_yield_*.csv"))
    if not files:
        raise RuntimeError("No Treasury yield curve cache files found")
    for file_path in files:
        reader = csv.DictReader(io.StringIO(file_path.read_text()))
        for row in reader:
            raw_date = row.get("Date")
            if not raw_date:
                continue
            date = datetime.strptime(raw_date, "%m/%d/%Y").date().isoformat()
            if field == "10y_minus_2y":
                ten = parse_float(row.get("10 Yr"))
                two = parse_float(row.get("2 Yr"))
                value = None if ten is None or two is None else ten - two
            else:
                value = parse_float(row.get(field))
            if value is not None:
                rows.append({"date": date, "value": value})
    rows.sort(key=lambda item: item["date"])
    return compact_history(rows, max_points), {
        "adapter": "treasury_yield_curve_cache",
        "cache_glob": "data/macro/raw/treasury_yield_*.csv",
        "field": field,
    }


def fetch_treasury_yield_curve_long(field: str, fred_series_id: str, *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    fred_rows, fred_meta = fetch_fred(fred_series_id, max_points=None)
    treasury_rows, treasury_meta = fetch_treasury_yield_curve(field, max_points=None)
    rows_by_date = {row["date"]: row["value"] for row in fred_rows}
    rows_by_date.update({row["date"]: row["value"] for row in treasury_rows})
    rows = [{"date": date, "value": value} for date, value in sorted(rows_by_date.items())]
    return compact_history(rows, max_points), {
        "adapter": "fred_long_history_plus_treasury_curve_cache",
        "components": {"fred": fred_meta, "treasury": treasury_meta},
        "gap": "Long history comes from FRED; overlapping 2016+ values are patched with local U.S. Treasury curve cache when available.",
    }


def fetch_treasury_real_yield_curve(field: str, *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows = []
    files = sorted(RAW_DIR.glob("treasury_real_yield_*.csv"))
    if not files:
        raise RuntimeError("No Treasury real yield curve cache files found")
    for file_path in files:
        reader = csv.DictReader(io.StringIO(file_path.read_text()))
        for row in reader:
            raw_date = row.get("Date")
            if not raw_date:
                continue
            value = parse_float(row.get(field))
            if value is not None:
                date = datetime.strptime(raw_date, "%m/%d/%Y").date().isoformat()
                rows.append({"date": date, "value": value})
    rows.sort(key=lambda item: item["date"])
    return compact_history(rows, max_points), {
        "adapter": "treasury_real_yield_curve_cache",
        "cache_glob": "data/macro/raw/treasury_real_yield_*.csv",
        "field": field,
    }


def fetch_treasury_real_yield_curve_long(field: str, fred_series_id: str, *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    fred_rows, fred_meta = fetch_fred(fred_series_id, max_points=None)
    treasury_rows, treasury_meta = fetch_treasury_real_yield_curve(field, max_points=None)
    rows_by_date = {row["date"]: row["value"] for row in fred_rows}
    rows_by_date.update({row["date"]: row["value"] for row in treasury_rows})
    rows = [{"date": date, "value": value} for date, value in sorted(rows_by_date.items())]
    return compact_history(rows, max_points), {
        "adapter": "fred_long_history_plus_treasury_real_yield_cache",
        "components": {"fred": fred_meta, "treasury": treasury_meta},
        "gap": "Long real-yield history comes from FRED DFII10; overlapping 2016+ values are patched with local U.S. Treasury TIPS curve cache when available.",
    }


def latest_value_on_or_before(rows: list[dict[str, Any]], date: str) -> float | None:
    latest = None
    for row in rows:
        if row["date"] <= date:
            latest = row.get("value")
        else:
            break
    return latest


def fetch_policy_rate_expectations_proxy(*, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    two_year_rows, two_year_meta = fetch_treasury_yield_curve_long("2 Yr", "DGS2", max_points=None)
    fed_funds_rows, fed_funds_meta = fetch_fred("FEDFUNDS", max_points=None)
    fed_funds_rows.sort(key=lambda item: item["date"])
    rows = []
    for row in two_year_rows:
        fed_funds = latest_value_on_or_before(fed_funds_rows, row["date"])
        if fed_funds is not None:
            rows.append({"date": row["date"], "value": row["value"] - fed_funds})
    return compact_history(rows, max_points), {
        "adapter": "derived_policy_expectations_proxy",
        "components": {"2Y Treasury": two_year_meta, "Fed Funds": fed_funds_meta},
        "gap": "Proxy only. Uses 2Y Treasury yield minus effective Fed Funds rate; it reflects market policy-cycle expectations, not CME FedWatch meeting probabilities.",
    }


def fetch_sp500_profit_valuation_proxy(*, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    sp500_rows, sp500_meta = fetch_fred("SP500", max_points=None)
    profits_rows, profits_meta = fetch_fred("CPATAX", max_points=None)
    sp500_rows.sort(key=lambda item: item["date"])
    rows = []
    for row in profits_rows:
        sp500 = latest_value_on_or_before(sp500_rows, row["date"])
        profit = row.get("value")
        if sp500 is not None and profit not in (None, 0):
            rows.append({"date": row["date"], "value": sp500 / profit})
    return compact_history(rows, max_points), {
        "adapter": "derived_sp500_profit_valuation_proxy",
        "components": {"S&P 500": sp500_meta, "NIPA after-tax corporate profits": profits_meta},
        "gap": "Proxy only. Uses S&P 500 index divided by NIPA after-tax corporate profits; not analyst-consensus forward P/E.",
    }


def fetch_fred_yoy(series_id: str, *, source_name: str, gap: str, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows, meta = fetch_fred(series_id, max_points=None)
    yoy_rows = calculate_yoy(rows)
    return compact_history(yoy_rows, max_points), {
        **meta,
        "adapter": f"{meta.get('adapter')}_yoy",
        "source_name": source_name,
        "operation": "yoy",
        "gap": gap,
    }


def fetch_treasury_rate_volatility_proxy(*, max_points: int | None, window: int = 60) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows, meta = fetch_treasury_yield_curve("10 Yr", max_points=None)
    vol_rows = []
    changes = []
    previous = None
    for row in rows:
        current = row.get("value")
        if previous is not None and current is not None:
            changes.append(current - previous)
            if len(changes) >= window:
                window_values = changes[-window:]
                mean = sum(window_values) / window
                variance = sum((value - mean) ** 2 for value in window_values) / (window - 1)
                vol_rows.append({"date": row["date"], "value": math.sqrt(variance) * math.sqrt(252)})
        previous = current
    return compact_history(vol_rows, max_points), {
        "adapter": "derived_treasury_10y_realized_volatility_proxy",
        "window_days": window,
        "components": {"10Y Treasury": meta},
        "gap": "Proxy only. Uses annualized 60-trading-day realized volatility of daily 10Y Treasury yield changes; not ICE/BofA MOVE.",
    }


def fetch_coingecko_coin(coin_id: str, *, vs_currency: str, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    url = (
        f"https://api.coingecko.com/api/v3/coins/{coin_id}/market_chart?"
        f"{urlencode({'vs_currency': vs_currency, 'days': 'max', 'interval': 'daily'})}"
    )
    payload = json.loads(run_curl(url, timeout=45))
    rows = []
    for ts, value in payload.get("prices", []):
        date = datetime.fromtimestamp(ts / 1000, tz=timezone.utc).date().isoformat()
        parsed = parse_float(value)
        if parsed is not None:
            rows.append({"date": date, "value": parsed})
    return compact_history(rows, max_points), {
        "adapter": "coingecko_market_chart",
        "url": url,
        "coin_id": coin_id,
        "vs_currency": vs_currency,
    }


def fetch_btc_dominance(*, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    # Free public endpoints usually expose latest dominance only. Preserve it as
    # a one-point series and keep the history gap visible.
    coingecko_url = "https://api.coingecko.com/api/v3/global"
    coinlore_url = "https://api.coinlore.net/api/global/"
    source_meta: dict[str, Any]
    try:
        cache_path = RAW_DIR / "coingecko_global.json"
        text, meta = fetch_text_live_first(coingecko_url, cache_path, timeout=30, reject_html=True)
        assert isinstance(text, str)
        payload = json.loads(text)
        value = payload.get("data", {}).get("market_cap_percentage", {}).get("btc")
        source_meta = {"adapter": "coingecko_global_latest_only" if meta.get("refreshed") else "coingecko_global_latest_only_cache_fallback", **meta}
    except Exception as exc:  # noqa: BLE001
        cache_path = RAW_DIR / "coinlore_global.json"
        text, meta = fetch_text_live_first(coinlore_url, cache_path, timeout=30, reject_html=True)
        assert isinstance(text, str)
        payload = json.loads(text)
        value = (payload[0] if payload else {}).get("btc_d")
        source_meta = {
            "adapter": "coinlore_global_latest_only" if meta.get("refreshed") else "coinlore_global_latest_only_cache_fallback",
            **meta,
            "fallback_from": coingecko_url,
            "fallback_error": str(exc),
        }
    parsed = parse_float(value)
    rows = []
    if parsed is not None:
        rows.append({"date": datetime.now(timezone.utc).date().isoformat(), "value": parsed})
    return compact_history(rows, max_points), {
        **source_meta,
        "gap": "Free endpoint provides latest dominance only; historical dominance needs derived market-cap history.",
    }


def fetch_stablecoin_supply(*, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    url = "https://stablecoins.llama.fi/stablecoincharts/all"
    cache_path = RAW_DIR / "defillama_stablecoincharts_all.json"
    text, meta = fetch_text_live_first(url, cache_path, timeout=45, reject_html=True)
    assert isinstance(text, str)
    payload = json.loads(text)
    source_meta = {"adapter": "defillama_stablecoincharts_all" if meta.get("refreshed") else "defillama_stablecoincharts_all_cache_fallback", **meta}
    rows = []
    for item in payload or []:
        ts = item.get("date")
        value = ((item.get("totalCirculatingUSD") or {}).get("peggedUSD"))
        parsed = parse_float(value)
        if ts and parsed is not None:
            rows.append({"date": datetime.fromtimestamp(int(ts), tz=timezone.utc).date().isoformat(), "value": parsed})
    rows.sort(key=lambda item: item["date"])
    return compact_history(rows, max_points), {
        **source_meta,
        "gap": "DeFiLlama all-stablecoin circulating USD-pegged supply history; methodology may revise historical chain/asset classification.",
    }


def fetch_defi_tvl(*, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    url = "https://api.llama.fi/v2/historicalChainTvl"
    cache_path = RAW_DIR / "defillama_tvl.json"
    text, meta = fetch_text_live_first(url, cache_path, timeout=60, reject_html=True)
    assert isinstance(text, str)
    payload = json.loads(text)
    source_meta = {"adapter": "defillama_historical_chain_tvl" if meta.get("refreshed") else "defillama_historical_chain_tvl_cache_fallback", **meta}
    rows = []
    for item in payload:
        ts = item.get("date")
        value = parse_float(item.get("tvl"))
        if ts and value is not None:
            rows.append({"date": datetime.fromtimestamp(int(ts), tz=timezone.utc).date().isoformat(), "value": value})
    return compact_history(rows, max_points), source_meta


def month_date(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "nat"}:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("/", "-"))
        return parsed.date().isoformat()
    except ValueError:
        pass
    if len(text) == 7 and text[4] == "-":
        return f"{text}-01"
    if len(text) == 6 and text.isdigit():
        return f"{text[:4]}-{text[4:]}-01"
    return None


def calculate_yoy(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_month = {row["date"][:7]: row["value"] for row in rows if row.get("value") is not None}
    yoy_rows = []
    for row in rows:
        previous_month = f"{int(row['date'][:4]) - 1:04d}-{row['date'][5:7]}"
        previous = by_month.get(previous_month)
        current = row.get("value")
        if previous not in (None, 0) and current is not None:
            yoy_rows.append({"date": row["date"], "value": (current / previous - 1) * 100})
    return yoy_rows


def calculate_mom_chain_yoy(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_month = {row["date"][:7]: row["value"] for row in rows if row.get("value") is not None}
    yoy_rows = []
    for row in rows:
        year = int(row["date"][:4])
        month = int(row["date"][5:7])
        factors = []
        for offset in range(12):
            index = year * 12 + month - offset - 1
            key = f"{index // 12:04d}-{index % 12 + 1:02d}"
            mom = by_month.get(key)
            if mom is None:
                break
            factors.append(1 + mom / 100)
        if len(factors) == 12:
            value = math.prod(factors) - 1
            yoy_rows.append({"date": row["date"], "value": value * 100})
    return yoy_rows


def fetch_macropage_xlsx(spec: dict[str, Any], *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for macropage xlsx caches; use the Codex bundled Python runtime or install openpyxl"
        ) from exc

    path = RAW_DIR / spec["file"]
    if not path.exists():
        raise RuntimeError(f"Missing macropage cache file: {path.relative_to(ROOT)}")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    sheet.reset_dimensions()
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows_iter)]
    date_index = headers.index("日期") if "日期" in headers else None
    if date_index is None:
        raise RuntimeError(f"No 日期 column in {path.relative_to(ROOT)}")

    operation = spec.get("operation")
    if operation == "average_columns_contains":
        contains = spec["contains"]
        value_indexes = [index for index, header in enumerate(headers) if contains in header]
        if not value_indexes:
            raise RuntimeError(f"No columns containing {contains!r} in {path.relative_to(ROOT)}")
    else:
        column = spec["column"]
        if column not in headers:
            raise RuntimeError(f"No column {column!r} in {path.relative_to(ROOT)}")
        value_indexes = [headers.index(column)]

    rows = []
    for raw_row in rows_iter:
        date = month_date(raw_row[date_index] if date_index < len(raw_row) else None)
        if not date:
            continue
        values = []
        for index in value_indexes:
            if index < len(raw_row):
                parsed = parse_float(raw_row[index])
                if parsed is not None:
                    values.append(parsed)
        if not values:
            continue
        rows.append({"date": date, "value": sum(values) / len(values)})
    rows.sort(key=lambda item: item["date"])

    if operation == "yoy":
        rows = calculate_yoy(rows)
    elif operation == "mom_chain_yoy":
        rows = calculate_mom_chain_yoy(rows)

    meta = {
        "adapter": "macropage_xlsx_cache",
        "cache_path": str(path.relative_to(ROOT)),
        "source_name": spec.get("source_name"),
        "operation": operation or "column_level",
        "column": spec.get("column"),
        "contains": spec.get("contains"),
        "source_url": "https://github.com/AFAN-LIFE/macropage",
        "gap": "Curated open-source cache; use official NBS/PBOC/Customs or AKShare for refresh automation.",
    }
    return compact_history(rows, max_points), meta


def fetch_macropage_csv(spec: dict[str, Any], *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    path = RAW_DIR / spec["file"]
    if not path.exists():
        raise RuntimeError(f"Missing macropage cache file: {path.relative_to(ROOT)}")
    reader = csv.DictReader(io.StringIO(path.read_text()))
    rows = []
    for row in reader:
        date = row.get(spec["date_column"])
        value = parse_float(row.get(spec["value_column"]))
        if date and value is not None:
            rows.append({"date": date, "value": value})
    rows.sort(key=lambda item: item["date"])
    return compact_history(rows, max_points), {
        "adapter": "macropage_csv_cache",
        "cache_path": str(path.relative_to(ROOT)),
        "source_name": spec.get("source_name"),
        "date_column": spec.get("date_column"),
        "value_column": spec.get("value_column"),
        "source_url": "https://github.com/AFAN-LIFE/macropage",
        "gap": "Curated open-source cache; use official ChinaBond or AKShare for refresh automation.",
    }


def fetch_jin10_report(spec: dict[str, Any], *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    cached_rows, cached_meta = fetch_jin10_report_cache(spec, max_points=max_points)
    official_rows, official_meta = fetch_ism_official_recent(spec, max_points=max_points)
    if cached_rows and official_rows:
        rows_by_date = {row["date"]: row["value"] for row in cached_rows}
        rows_by_date.update({row["date"]: row["value"] for row in official_rows})
        rows = [{"date": date, "value": value} for date, value in sorted(rows_by_date.items())]
        return compact_history(rows, max_points), {
            "adapter": "jin10_cache_plus_ism_official_recent",
            "components": {"history_seed": cached_meta, "official_recent": official_meta},
            "gap": "Long history seeded from Jin10 public calendar cache; latest 12 months patched from official ISM report page.",
        }
    if cached_rows:
        return cached_rows, cached_meta

    url = "https://datacenter-api.jin10.com/reports/list_v2"
    headers = [
        "-H",
        "user-agent: Mozilla/5.0",
        "-H",
        "x-app-id: rU6QIu7JHe2gOUeR",
        "-H",
        "x-csrf-token: x-csrf-token",
        "-H",
        "x-version: 1.0.0",
    ]
    max_date = ""
    raw_rows = []
    while True:
        query = urlencode({"max_date": max_date, "category": "ec", "attr_id": spec["attr_id"], "_": int(time.time() * 1000)})
        args = ["curl", "-sS", "-L", "--max-time", "30", *headers, f"{url}?{query}"]
        completed = subprocess.run(
            args,
            cwd=ROOT,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )
        if completed.returncode != 0:
            raise RuntimeError(completed.stderr.strip() or f"curl exited {completed.returncode}")
        payload = json.loads(completed.stdout)
        values = ((payload.get("data") or {}).get("values") or [])
        if not values:
            break
        raw_rows.extend(values)
        last_date = values[-1][0]
        previous_day = datetime.strptime(last_date, "%Y-%m-%d").date().toordinal() - 1
        max_date = datetime.fromordinal(previous_day).date().isoformat()
        if len(values) < 20:
            break
    rows = []
    for item in raw_rows:
        date = item[0]
        value = parse_float(item[1])
        if date and value is not None:
            rows.append({"date": date, "value": value})
    rows.sort(key=lambda item: item["date"])
    return compact_history(rows, max_points), {
        "adapter": "jin10_report_list_v2",
        "source_name": f"Jin10 datacenter / {spec['symbol']}",
        "source_url": spec.get("source_url"),
        "attr_id": spec["attr_id"],
        "gap": "Third-party public economic-calendar feed; replace with NBS/AKShare official-source adapter when available.",
    }


def fetch_ism_official_recent(spec: dict[str, Any], *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    cache_file = spec.get("official_recent_cache")
    url = spec.get("official_recent_url")
    if not cache_file or not url:
        return [], {}
    cache_path = RAW_DIR / cache_file
    text, meta = fetch_text_live_first(url, cache_path, timeout=45)
    assert isinstance(text, str)
    source_meta = {
        "adapter": "ism_official_recent" if meta.get("refreshed") else "ism_official_recent_cache_fallback",
        "source_name": spec.get("official_recent_source"),
        "source_url": url,
        **meta,
    }

    start_marker = spec.get("official_recent_start")
    end_marker = spec.get("official_recent_end")
    start = text.find(start_marker) if start_marker else 0
    if start == -1:
        return [], {**source_meta, "error": "official recent start marker not found"}
    end = text.find(end_marker, start) if end_marker else -1
    section = text[start : end if end != -1 else start + 8000]
    month_pattern = "|".join(MONTH_ABBR_TO_NUMBER)
    pattern = re.compile(
        rf"<(?:th|td)[^>]*>\s*({month_pattern})\s+(\d{{4}})\s*</(?:th|td)>\s*<td[^>]*>\s*([0-9]+(?:\.[0-9]+)?)\s*</td>",
        re.IGNORECASE | re.DOTALL,
    )
    rows = []
    for match in pattern.finditer(section):
        month, year, value_text = match.groups()
        value = parse_float(value_text)
        if value is None:
            continue
        rows.append({"date": f"{year}-{MONTH_ABBR_TO_NUMBER[month[:3].title()]}-01", "value": value})
    rows.sort(key=lambda item: item["date"])
    return compact_history(rows, max_points), {
        **source_meta,
        "gap": "Official ISM report page exposes the latest 12 months only; use with long-history seed.",
    }


def fetch_jin10_report_cache(spec: dict[str, Any], *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows_by_date: dict[str, float] = {}
    files_used = []
    all_js_file = spec.get("all_js_file")
    all_js_path = RAW_DIR / all_js_file if all_js_file else None
    if all_js_path and all_js_path.exists():
        text = all_js_path.read_text()
        prefix = "var dataCenter_data = "
        if text.startswith(prefix):
            payload = json.loads(text[len(prefix):].rstrip().rstrip(";"))
            for item in payload.get("list", []):
                raw_date = str(item.get("date") or "")
                values = ((item.get("datas") or {}).get(spec["symbol"]) or [])
                value = parse_float(values[1] if len(values) > 1 else None)
                if len(raw_date) == 8 and value is not None:
                    date = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:]}"
                    rows_by_date[date] = value
            files_used.append(str(all_js_path.relative_to(ROOT)))

    latest_path = RAW_DIR / f"jin10_attr_{spec['attr_id']}_latest.json"
    if latest_path.exists():
        payload = json.loads(latest_path.read_text())
        for item in ((payload.get("data") or {}).get("values") or []):
            date = item[0]
            value = parse_float(item[1])
            if date and value is not None:
                rows_by_date[date] = value
        files_used.append(str(latest_path.relative_to(ROOT)))

    rows = [{"date": date, "value": value} for date, value in sorted(rows_by_date.items())]
    return compact_history(rows, max_points), {
        "adapter": "jin10_report_cache",
        "source_name": f"Jin10 datacenter / {spec['symbol']}",
        "source_url": spec.get("source_url"),
        "attr_id": spec["attr_id"],
        "cache_files": files_used,
        "gap": "Third-party public economic-calendar cache; replace with NBS/AKShare official-source adapter when available.",
    }


def hyperliquid_post(body: dict[str, Any]) -> Any:
    text = run_curl("https://api.hyperliquid.xyz/info", method="POST", data=body, timeout=45)
    return json.loads(text)


def fetch_hyperliquid_candles(coin: str, *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    cache_path = RAW_DIR / f"hyperliquid_{coin.lower()}_1d_candles.json"
    end_ms = int(time.time() * 1000)
    start_ms = end_ms - 10 * 365 * 24 * 60 * 60 * 1000
    body = {
        "type": "candleSnapshot",
        "req": {
            "coin": coin,
            "interval": "1d",
            "startTime": start_ms,
            "endTime": end_ms,
        },
    }
    try:
        payload = hyperliquid_post(body)
        cache_path.write_text(json.dumps(payload, ensure_ascii=False) + "\n")
        source_meta = {
            "adapter": "hyperliquid_candle_snapshot",
            "cache_path": str(cache_path.relative_to(ROOT)),
            "refreshed": True,
            "coin": coin,
        }
    except Exception as exc:  # noqa: BLE001
        if not cache_path.exists():
            raise
        payload = json.loads(cache_path.read_text())
        source_meta = {
            "adapter": "hyperliquid_candle_snapshot_cache_fallback",
            "cache_path": str(cache_path.relative_to(ROOT)),
            "refreshed": False,
            "cache_fallback": True,
            "live_error": str(exc),
            "coin": coin,
        }
    rows = []
    for item in payload or []:
        ts = item.get("t") or item.get("T")
        close = parse_float(item.get("c"))
        if ts and close is not None:
            rows.append({"date": datetime.fromtimestamp(int(ts) / 1000, tz=timezone.utc).date().isoformat(), "value": close})
    rows.sort(key=lambda item: item["date"])
    return compact_history(rows, max_points), {
        **source_meta,
        "gap": "Hyperliquid perpetual daily close; venue-level perp price, not a consolidated spot index.",
    }


def fetch_hyperliquid_eth_btc(*, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    btc_rows, btc_meta = fetch_hyperliquid_candles("BTC", max_points=None)
    eth_rows, eth_meta = fetch_hyperliquid_candles("ETH", max_points=None)
    btc_by_date = {row["date"]: row["value"] for row in btc_rows}
    rows = []
    for row in eth_rows:
        btc = btc_by_date.get(row["date"])
        if btc:
            rows.append({"date": row["date"], "value": row["value"] / btc})
    return compact_history(rows, max_points), {
        "adapter": "hyperliquid_candle_snapshot_cache_derived",
        "legs": {"ETH": eth_meta, "BTC": btc_meta},
        "gap": "Derived from Hyperliquid ETH and BTC perpetual daily closes; not a consolidated spot ETH/BTC index.",
    }


def fetch_hyperliquid_funding(*, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows_by_date: dict[str, list[float]] = {}
    end_ms = int(time.time() * 1000)
    start_ms = end_ms - 365 * 24 * 60 * 60 * 1000
    errors = []
    for coin in HYPERLIQUID_COINS:
        try:
            cursor = start_ms
            for _ in range(40):
                data = hyperliquid_post({"type": "fundingHistory", "coin": coin, "startTime": cursor, "endTime": end_ms})
                if not data:
                    break
                last_ts = None
                for item in data:
                    ts = item.get("time")
                    rate = parse_float(item.get("fundingRate"))
                    if ts and rate is not None:
                        last_ts = int(ts)
                        date = datetime.fromtimestamp(last_ts / 1000, tz=timezone.utc).date().isoformat()
                        rows_by_date.setdefault(date, []).append(rate)
                if last_ts is None or last_ts <= cursor or last_ts >= end_ms:
                    break
                cursor = last_ts + 1
                if len(data) < 500:
                    break
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{coin}: {exc}")
    rows = [
        {"date": date, "value": sum(values) / len(values)}
        for date, values in sorted(rows_by_date.items())
        if values
    ]
    if not rows:
        cached_rows, cached_meta = fetch_hyperliquid_cache_series("funding", max_points=max_points)
        if cached_rows:
            return cached_rows, {**cached_meta, "live_errors": errors}
    return compact_history(rows, max_points), {
        "adapter": "hyperliquid_funding_history",
        "coins": HYPERLIQUID_COINS,
        "gap": "Venue-level BTC/ETH/SOL average, not whole-market aggregate.",
        "errors": errors,
    }


def fetch_hyperliquid_open_interest(*, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    cached_rows, cached_meta = fetch_hyperliquid_cache_series("open_interest_usd", max_points=max_points)
    if cached_rows:
        return cached_rows, cached_meta

    data = hyperliquid_post({"type": "metaAndAssetCtxs"})
    universe = data[0].get("universe", []) if isinstance(data, list) and data else []
    ctxs = data[1] if isinstance(data, list) and len(data) > 1 else []
    total = 0.0
    matched = []
    for meta, ctx in zip(universe, ctxs):
        coin = meta.get("name")
        if coin not in HYPERLIQUID_COINS:
            continue
        oi = parse_float(ctx.get("openInterest"))
        mark = parse_float(ctx.get("markPx"))
        if oi is not None and mark is not None:
            total += oi * mark
            matched.append(coin)
    rows = [{"date": datetime.now(timezone.utc).date().isoformat(), "value": total}] if total else []
    return compact_history(rows, max_points), {
        "adapter": "hyperliquid_meta_and_asset_contexts",
        "coins": matched,
        "gap": "Latest venue-level notional OI for BTC/ETH/SOL only, not whole-market aggregate.",
    }


def fetch_hyperliquid_cache_series(metric: str, *, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    latest_by_coin_date: dict[str, dict[str, dict[str, Any]]] = {}
    files_used = []
    skipped = 0
    for coin in HYPERLIQUID_COINS:
        path = CRYPTO_OI_DIR / f"{coin.lower()}.jsonl"
        if not path.exists():
            continue
        files_used.append(str(path.relative_to(ROOT)))
        for line in path.read_text().splitlines():
            if not line.strip():
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                skipped += 1
                continue
            timestamp = parse_float(item.get("timestamp"))
            iso_time = item.get("isoTime")
            if timestamp is not None:
                date = datetime.fromtimestamp(timestamp / 1000, tz=timezone.utc).date().isoformat()
            elif iso_time:
                date = str(iso_time)[:10]
            else:
                skipped += 1
                continue
            latest = latest_by_coin_date.setdefault(coin, {}).get(date)
            item_ts = parse_float(item.get("timestamp")) or 0
            latest_ts = parse_float(latest.get("timestamp")) or 0 if latest else 0
            if latest is None or item_ts > latest_ts:
                latest_by_coin_date[coin][date] = item

    values_by_date: dict[str, list[float]] = {}
    for rows_by_date in latest_by_coin_date.values():
        for date, item in rows_by_date.items():
            if metric == "open_interest_usd":
                oi = parse_float(item.get("openInterestBase"))
                mark = parse_float(item.get("markPx"))
                value = None if oi is None or mark is None else oi * mark
            elif metric == "funding":
                value = parse_float(item.get("funding"))
            else:
                value = None
            if value is not None:
                values_by_date.setdefault(date, []).append(value)

    if metric == "open_interest_usd":
        rows = [{"date": date, "value": sum(values)} for date, values in sorted(values_by_date.items()) if values]
        gap = "Daily latest cached Hyperliquid notional OI for BTC/ETH/SOL only, not whole-market aggregate."
    else:
        rows = [
            {"date": date, "value": sum(values) / len(values)}
            for date, values in sorted(values_by_date.items())
            if values
        ]
        gap = "Daily latest cached average funding for BTC/ETH/SOL only, not whole-market aggregate."

    return compact_history(rows, max_points), {
        "adapter": "hyperliquid_local_oi_history_cache",
        "metric": metric,
        "coins": HYPERLIQUID_COINS,
        "cache_files": files_used,
        "skipped_bad_lines": skipped,
        "gap": gap,
    }


def fetch_global_manufacturing_pmi_proxy(*, max_points: int | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    component_specs = [
        ("us_ism_manufacturing_pmi", 0.45, lambda: fetch_jin10_report(JIN10_REPORTS["us_ism_manufacturing_pmi"], max_points=None)),
        ("cn_official_manufacturing_pmi", 0.35, lambda: fetch_macropage_xlsx(MACROPAGE_XLSX_SERIES["cn_official_manufacturing_pmi"], max_points=None)),
        ("cn_caixin_manufacturing_pmi", 0.20, lambda: fetch_jin10_report(JIN10_REPORTS["cn_caixin_manufacturing_pmi"], max_points=None)),
    ]
    values_by_month: dict[str, dict[str, float]] = {}
    meta_by_component: dict[str, Any] = {}
    weights = {}
    for component_id, weight, fetcher in component_specs:
        component_rows, component_meta = fetcher()
        weights[component_id] = weight
        meta_by_component[component_id] = component_meta
        for row in component_rows:
            value = row.get("value")
            date = row.get("date")
            if date and value is not None:
                values_by_month.setdefault(date[:7], {})[component_id] = value

    rows = []
    for month, component_values in sorted(values_by_month.items()):
        available_weight = sum(weights[key] for key in component_values)
        if not available_weight:
            continue
        value = sum(component_values[key] * weights[key] for key in component_values) / available_weight
        rows.append({"date": f"{month}-01", "value": value})

    return compact_history(rows, max_points), {
        "adapter": "derived_weighted_pmi_proxy",
        "weights": weights,
        "components": meta_by_component,
        "gap": "Proxy only. This is not the official J.P. Morgan/S&P Global Manufacturing PMI; weights are normalized across available US/China components.",
    }


def fetch_indicator(indicator: dict[str, Any], *, max_points: int | None) -> dict[str, Any]:
    indicator_id = indicator["id"]
    history_path = HISTORY_DIR / f"{indicator_id}.json"
    result = {
        "id": indicator_id,
        "priority": indicator["priority"],
        "label": indicator["label"],
        "region": indicator["region"],
        "theme": indicator["theme"],
        "configured_source": indicator.get("source"),
        "configured_source_url": indicator.get("source_url"),
        "display_frequency": indicator.get("display_frequency"),
        "transform": indicator.get("transform"),
        "availability": indicator.get("availability"),
        "generated_at": now_iso(),
    }
    try:
        series_id = indicator.get("series_id")
        rows: list[dict[str, Any]]
        meta: dict[str, Any]
        if indicator_id in TREASURY_YIELD_INDICATORS:
            field, fred_series_id = TREASURY_YIELD_INDICATORS[indicator_id]
            rows, meta = fetch_treasury_yield_curve_long(field, fred_series_id, max_points=max_points)
        elif indicator_id in TREASURY_REAL_YIELD_INDICATORS:
            field, fred_series_id = TREASURY_REAL_YIELD_INDICATORS[indicator_id]
            rows, meta = fetch_treasury_real_yield_curve_long(field, fred_series_id, max_points=max_points)
        elif indicator_id == "us_policy_rate_expectations":
            rows, meta = fetch_policy_rate_expectations_proxy(max_points=max_points)
        elif indicator_id in PBOC_STATS_SERIES:
            rows, meta = fetch_pboc_stats_series(indicator_id, PBOC_STATS_SERIES[indicator_id], max_points=max_points)
        elif indicator_id in NBS_RELEASE_SERIES:
            rows, meta = fetch_nbs_release_series(indicator_id, NBS_RELEASE_SERIES[indicator_id], max_points=max_points)
        elif indicator_id in EASTMONEY_SERIES:
            rows, meta = fetch_eastmoney_series(EASTMONEY_SERIES[indicator_id], max_points=max_points)
        elif indicator_id in EASTMONEY_HOUSE_PRICE_SERIES:
            rows, meta = fetch_eastmoney_house_price_average(EASTMONEY_HOUSE_PRICE_SERIES[indicator_id], max_points=max_points)
        elif indicator_id in MOFCOM_SOCIAL_FINANCING_FLOW:
            rows, meta = fetch_mofcom_social_financing_flow(MOFCOM_SOCIAL_FINANCING_FLOW[indicator_id], max_points=max_points)
        elif indicator_id == "us_sp500_forward_pe":
            rows, meta = fetch_sp500_profit_valuation_proxy(max_points=max_points)
        elif indicator_id == "us_sp500_eps_growth":
            rows, meta = fetch_fred_yoy(
                "CPATAX",
                source_name="FRED / BEA NIPA corporate profits after tax",
                gap="Proxy only. Uses NIPA after-tax corporate profits YoY; not analyst-consensus S&P 500 EPS growth.",
                max_points=max_points,
            )
        elif indicator_id == "de_manufacturing_pmi":
            rows, meta = fetch_fred_yoy(
                "DEUPROINDMISMEI",
                source_name="FRED/OECD Germany production of total industry",
                gap="Proxy only. Uses Germany industrial production YoY; not S&P Global Germany Manufacturing PMI.",
                max_points=max_points,
            )
        elif indicator_id == "xasset_move_index":
            rows, meta = fetch_treasury_rate_volatility_proxy(max_points=max_points)
        elif series_id in FRED_SERIES_IDS:
            rows, meta = fetch_fred(series_id, max_points=max_points)
        elif indicator_id in STOOQ_SYMBOLS:
            rows, meta = fetch_stooq(STOOQ_SYMBOLS[indicator_id], max_points=max_points)
        elif indicator_id in EUROSTAT_SERIES:
            rows, meta = fetch_eurostat_jsonstat(EUROSTAT_SERIES[indicator_id], max_points=max_points)
        elif indicator_id in NASDAQ_SYMBOLS:
            symbol, assetclass = NASDAQ_SYMBOLS[indicator_id]
            rows, meta = fetch_nasdaq_history(symbol, assetclass, max_points=max_points)
        elif indicator_id == "xasset_btc_price_trend":
            rows, meta = fetch_hyperliquid_candles("BTC", max_points=max_points)
        elif indicator_id == "xasset_eth_btc":
            rows, meta = fetch_hyperliquid_eth_btc(max_points=max_points)
        elif indicator_id == "xasset_btc_dominance":
            rows, meta = fetch_btc_dominance(max_points=max_points)
        elif indicator_id == "xasset_stablecoin_supply":
            rows, meta = fetch_stablecoin_supply(max_points=max_points)
        elif indicator_id == "xasset_defi_tvl":
            rows, meta = fetch_defi_tvl(max_points=max_points)
        elif indicator_id == "xasset_crypto_funding_rate":
            rows, meta = fetch_hyperliquid_funding(max_points=max_points)
        elif indicator_id == "xasset_crypto_open_interest":
            rows, meta = fetch_hyperliquid_open_interest(max_points=max_points)
        elif indicator_id in MACROPAGE_XLSX_SERIES:
            rows, meta = fetch_macropage_xlsx(MACROPAGE_XLSX_SERIES[indicator_id], max_points=max_points)
        elif indicator_id in MACROPAGE_CSV_SERIES:
            rows, meta = fetch_macropage_csv(MACROPAGE_CSV_SERIES[indicator_id], max_points=max_points)
        elif indicator_id in JIN10_REPORTS:
            rows, meta = fetch_jin10_report(JIN10_REPORTS[indicator_id], max_points=max_points)
        elif indicator_id == "global_manufacturing_pmi":
            rows, meta = fetch_global_manufacturing_pmi_proxy(max_points=max_points)
        else:
            history_path = HISTORY_DIR / f"{indicator_id}.json"
            history_path.write_text("[]\n")
            return {
                **result,
                "status": "missing_adapter",
                "points": 0,
                "first_date": None,
                "latest_date": None,
                "latest_value": None,
                "history_path": str(history_path.relative_to(ROOT)),
                "gap": "No adapter implemented yet for this source/indicator.",
                "fallback": indicator.get("fallback"),
            }
        history_path = HISTORY_DIR / f"{indicator_id}.json"
        summary = summarize_history(rows)
        if summary.get("status") == "missing" and history_path.exists():
            try:
                cached_rows = json.loads(history_path.read_text())
            except json.JSONDecodeError:
                cached_rows = []
            if cached_rows:
                cached_summary = summarize_history(cached_rows)
                return {
                    **result,
                    **cached_summary,
                    "status": "stale_cache",
                    "history_path": str(history_path.relative_to(ROOT)),
                    "adapter_meta": {
                        "adapter": "existing_history_cache_fallback",
                        "cache_path": str(history_path.relative_to(ROOT)),
                        "refreshed": False,
                        "cache_fallback": True,
                        "live_error": "Latest fetch parsed no valid rows; preserved previous history.",
                        "attempted_adapter_meta": meta,
                    },
                    "fallback": indicator.get("fallback"),
                    "note": indicator.get("note"),
                }
        history_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n")
        return {
            **result,
            **summary,
            "history_path": str(history_path.relative_to(ROOT)),
            "adapter_meta": meta,
            "fallback": indicator.get("fallback"),
            "note": indicator.get("note"),
        }
    except Exception as exc:  # noqa: BLE001
        cached_rows: list[dict[str, Any]] = []
        if history_path.exists():
            try:
                cached_rows = json.loads(history_path.read_text())
            except json.JSONDecodeError:
                cached_rows = []
        if cached_rows:
            summary = summarize_history(cached_rows)
            return {
                **result,
                **summary,
                "status": "stale_cache",
                "history_path": str(history_path.relative_to(ROOT)),
                "adapter_meta": {
                    "adapter": "existing_history_cache_fallback",
                    "cache_path": str(history_path.relative_to(ROOT)),
                    "refreshed": False,
                    "cache_fallback": True,
                    "live_error": str(exc),
                },
                "error": str(exc),
                "fallback": indicator.get("fallback"),
                "note": indicator.get("note"),
            }
        history_path.write_text("[]\n")
        return {
            **result,
            "status": "error",
            "points": 0,
            "first_date": None,
            "latest_date": None,
            "latest_value": None,
            "history_path": str(history_path.relative_to(ROOT)),
            "error": str(exc),
            "fallback": indicator.get("fallback"),
            "note": indicator.get("note"),
        }


def write_csv_summary(rows: list[dict[str, Any]], path: Path) -> None:
    fields = [
        "priority",
        "id",
        "label",
        "region",
        "theme",
        "status",
        "points",
        "first_date",
        "latest_date",
        "latest_value",
        "display_frequency",
        "availability",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field) for field in fields})


def main() -> int:
    global OUT_DIR, HISTORY_DIR

    parser = argparse.ArgumentParser(description="Fetch macro indicator histories.")
    parser.add_argument("--config", type=Path, default=CONFIG_PATH)
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR)
    parser.add_argument("--max-points", type=int, default=2500, help="Keep only the latest N history points per indicator.")
    parser.add_argument("--limit", type=int, default=None, help="Fetch only the first N indicators for smoke tests.")
    args = parser.parse_args()

    OUT_DIR = args.out_dir
    HISTORY_DIR = OUT_DIR / "history"
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)

    config = yaml.safe_load(args.config.read_text())
    indicators = config["indicators"]
    if args.limit:
        indicators = indicators[: args.limit]

    summaries = []
    for indicator in indicators:
        print(f"[{indicator['priority']:02d}/70] {indicator['id']}", file=sys.stderr)
        summaries.append(fetch_indicator(indicator, max_points=args.max_points))

    payload = {
        "generated_at": now_iso(),
        "config": str(args.config.relative_to(ROOT) if args.config.is_relative_to(ROOT) else args.config),
        "count": len(summaries),
        "status_counts": {},
        "indicators": summaries,
    }
    for row in summaries:
        payload["status_counts"][row["status"]] = payload["status_counts"].get(row["status"], 0) + 1

    latest_path = OUT_DIR / "latest.json"
    latest_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n")
    write_csv_summary(summaries, OUT_DIR / "latest-summary.csv")
    print(json.dumps(payload["status_counts"], ensure_ascii=False, indent=2))
    print(f"Wrote {latest_path.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
